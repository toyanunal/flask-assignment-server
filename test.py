from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import padding
from cryptography.hazmat.backends import default_backend
import os

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection

def encrypt_data(data, key):
    # Pad the data
    padder = padding.PKCS7(algorithms.AES.block_size).padder()
    padded_data = padder.update(data.encode()) + padder.finalize()
    
    # Encrypt the data
    iv = os.urandom(16)
    cipher = Cipher(algorithms.AES(key), modes.CFB(iv), backend=default_backend())
    encryptor = cipher.encryptor()
    encrypted_data = encryptor.update(padded_data) + encryptor.finalize()
    
    return iv + encrypted_data

def decrypt_data(encrypted_data, key):
    # Extract the IV
    iv = encrypted_data[:16]
    encrypted_data = encrypted_data[16:]
    
    # Decrypt the data
    cipher = Cipher(algorithms.AES(key), modes.CFB(iv), backend=default_backend())
    decryptor = cipher.decryptor()
    padded_data = decryptor.update(encrypted_data) + decryptor.finalize()
    
    # Unpad the data
    unpadder = padding.PKCS7(algorithms.AES.block_size).unpadder()
    data = unpadder.update(padded_data) + unpadder.finalize()
    
    return data.decode()

# Function to add a hidden worksheet and store encrypted data
def add_hidden_worksheet(workbook, data, key):
    # Encrypt the data
    encrypted_data = encrypt_data(data, key)

    # Add a hidden worksheet
    hidden_ws = workbook.create_sheet(title="HiddenData")
    hidden_ws.sheet_state = 'hidden'

    # Store the encrypted data in the hidden worksheet
    for i in range(len(encrypted_data)):
        hidden_ws.cell(row=1, column=i + 1, value=encrypted_data[i])

# Function to retrieve and decrypt data from a hidden worksheet
def retrieve_hidden_data(workbook, key):
    hidden_ws = workbook["HiddenData"]

    # Retrieve the encrypted data
    encrypted_data = bytes([hidden_ws.cell(row=1, column=i + 1).value for i in range(hidden_ws.max_column)])

    # Decrypt the data
    data = decrypt_data(encrypted_data, key)

    return data

# Example usage
key = os.urandom(32)  # AES-256 key
ext_user_username = 'e162798'
semester_info = '2023-2024SPRING'
data = f"{ext_user_username},{semester_info}"

# Create a new workbook and add the hidden worksheet
workbook = Workbook()
add_hidden_worksheet(workbook, data, key)

# Save the workbook
workbook.save('example.xlsx')

# Load the workbook and retrieve the hidden data
workbook = openpyxl.load_workbook('example.xlsx')
retrieved_data = retrieve_hidden_data(workbook, key)
print("Retrieved data:", retrieved_data)
